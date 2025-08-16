import os
import pytz
import json
import requests
from datetime import datetime
from pymongo import MongoClient
from base64 import b64encode
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import google.generativeai as genai

# === Configuration ===
sonar_host = os.getenv("SONAR_HOST_URL")
sonar_token = os.getenv("SONAR_AUTH_TOKEN")
project_key = os.getenv("SONAR_PROJECT_KEY")
mongo_uri = os.getenv("MONGO_URI")
mongo_db = os.getenv("MONGO_DB")
mongo_collection = os.getenv("MONGO_COLLECTION")
sonar_json_path = os.getenv("SONAR_JSON")
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
output_excel_path = os.path.join(os.getenv("WORKSPACE", "/tmp"), "ai_suggestions_report.xlsx")
ist = pytz.timezone('Asia/Kolkata')
timestamp = datetime.now(ist).strftime("%Y-%m-%d %H:%M:%S")
# === Auth Header ===
auth_header = {
    "Authorization": f"Basic {b64encode(f'{sonar_token}:'.encode()).decode()}"
}

# === Read SonarQube JSON ===
with open(sonar_json_path, 'r') as f:
    sonar_data = json.load(f)
    metrics = sonar_data.get("component", {}).get("measures", [])

# === Unsafe Metric Conditions ===
unsafe_conditions = {
    "coverage": lambda x: float(x) < 80,
    "bugs": lambda x: int(x) > 0,
    "vulnerabilities": lambda x: int(x) > 0,
    "code_smells": lambda x: int(x) > 0,
    "duplicated_lines_density": lambda x: float(x) > 10,
    "security_rating": lambda x: str(x) not in ["1.0", "2.0"],
    "reliability_rating": lambda x: str(x) not in ["1.0", "2.0"],
    "sqale_rating": lambda x: str(x) not in ["1.0", "2.0"],
    "alert_status": lambda x: x.upper() == "ERROR",
}

# === Gemini model setup ===
model = genai.GenerativeModel("models/gemini-2.0-flash-lite")

def concise_gemini_suggestion(prompt):
    try:
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        print(f"⚠️ Gemini error: {e}")
        return "Manual review required. Suggestion unavailable."

# === Prepare Flagged Metrics ===
flagged_metrics = []
for item in metrics:
    metric = item.get("metric")
    value = item.get("value")
    if metric in unsafe_conditions and unsafe_conditions[metric](value):
        flagged_metrics.append({"metric": metric, "value": value})

# === Filter & Prompt for Metric Suggestions ===
metric_suggestions = []
for item in flagged_metrics:
    if item["metric"] == "code_smells":
        continue
    prompt = (
        f"Metric: {item['metric']}, Value: {item['value']}\n"
        "Generate 3-4 clear, short, human-friendly action steps (each on a new line). "
        "Skip markdown and no AI roleplay."
    )
    suggestion = concise_gemini_suggestion(prompt)
    metric_suggestions.append({
        "type": "Project Metric",
        "metric": item["metric"],
        "value": item["value"],
        "line": "N/A",
        "suggestion": suggestion
    })

# === Fetch and Suggest for Code-Level Issues ===
issue_types = ["BUG", "VULNERABILITY", "CODE_SMELL"]
code_issues = []

for issue_type in issue_types:
    try:
        response = requests.get(
            f"{sonar_host}/api/issues/search",
            params={"componentKeys": project_key, "types": issue_type, "statuses": "OPEN,REOPENED", "ps": "100"},
            headers=auth_header,
            timeout=10
        )
        response.raise_for_status()
        code_issues.extend(response.json().get("issues", []))
    except Exception as e:
        print(f"❌ Failed to fetch {issue_type.lower()}s: {e}")

for issue in code_issues:
    prompt = (
        f"Type: {issue.get('type')}, Rule: {issue.get('rule')}, Message: {issue.get('message')}\n"
        "Give 3-4 short actionable steps to fix this issue. One per line. Avoid markdown and intro text."
    )
    suggestion = concise_gemini_suggestion(prompt)
    metric_suggestions.append({
        "type": "Code Issue",
        "metric": issue.get("type"),
        "value": issue.get("message"),
        "line": issue.get("line", "N/A"),
        "suggestion": suggestion
    })

# === MongoDB Save ===
document = {
    "project_key": project_key,
    "generated_at": timestamp,
    "flagged_issues": metric_suggestions
}

try:
    client = MongoClient(f"mongodb://{mongo_uri}", serverSelectionTimeoutMS=5000)
    db = client[mongo_db]
    col = db[mongo_collection]
    col.insert_one(document)
except Exception as e:
    print(f"❌ MongoDB error: {e}")

# === Excel Output with Styling ===
wb = Workbook()
ws = wb.active
ws.title = "AI Suggestions"
headers = ["Type", "Metric", "Value", "Line", "AI Suggestion"]
ws.append(headers)

header_font = Font(bold=True)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Apply header formatting
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Append data rows
for item in metric_suggestions:
    row = [item["type"], item["metric"], item["value"], item["line"], item["suggestion"]]
    ws.append(row)

# Format all cells
for row in ws.iter_rows(min_row=2, max_col=5):
    for cell in row:
        cell.alignment = wrap_alignment
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border

# Auto-adjust column width
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_length + 5, 60)

wb.save(output_excel_path)
print(f"✅ Excel report saved to: {output_excel_path}")
